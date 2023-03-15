from openpyxl import load_workbook, Workbook
import os
#import map_gen

not_exists = False

if not os.path.exists("location"):
    wb = Workbook()
    not_exists = True
else:
    wb = load_workbook(filename='data.xlsx')

ws = wb.active

if not_exists:
    ws['F1'] = 1
    ws['A1'] = 'широта'
    ws['B1'] = 'долгота'
    ws['C1'] = 'имя'
    ws['D1'] = 'время отправки'
    ws['E1'] = 'chat_id'

# grab the active worksheet

curr_row = ws.cell(row=1, column=6).value


def add_data(list_data):
    global curr_row
    already_in_base = False
    found_index = 0
    for info in range(2, curr_row + 1):
        print(ws['E{}'.format(info)].value)
        print(list_data[4])
        if ws['E{}'.format(info)].value == list_data[4]:
            already_in_base = True
            found_index = info
            print(found_index)
            break
    if not already_in_base:
        curr_row += 1
        ws.append(list_data)
        ws['D{}'.format(curr_row)].number_format = 'm/d/yy h:mm'
        ws['F1'] = curr_row
    else:
        ws['A{}'.format(found_index)] = list_data[0]
        ws['B{}'.format(found_index)] = list_data[1]
        ws['C{}'.format(found_index)] = list_data[2]
        ws['D{}'.format(found_index)] = list_data[3]
    wb.save("data.xlsx")


def prepare_data():
    global curr_row
    data_list = []
    for info in range(2, curr_row + 1):
        data_list.append([ws['A{}'.format(info)].value,
                          ws['B{}'.format(info)].value,
                          ws['C{}'.format(info)].value,
                          ws['D{}'.format(info)].value])
    return data_list
